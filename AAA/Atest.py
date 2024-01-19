from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def write_data_to_existing_sheet(file_path, sheet_name, data):
    # Load the existing workbook
    workbook = load_workbook(file_path)

    # Check if the sheet exists, if not, create it
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    # Select the sheet
    worksheet = workbook[sheet_name]

    # Clear existing data in the sheet
    worksheet.delete_rows(1, worksheet.max_row)

    # Write column headers
    header_row = 1
    for col_index, col_name in enumerate(data[0], start=1):
        worksheet.cell(row=header_row, column=col_index, value=col_name)

    # Write new data to the sheet
    for row_index, row_data in enumerate(data, start=2):
        for col_index, col_value in enumerate(row_data.values(), start=1):
            worksheet.cell(row=row_index, column=col_index, value=col_value)

    # Auto size columns based on content
    for column in worksheet.columns:
        max_length = 0
        column = [col for col in column]
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Save the changes to the existing workbook
    workbook.save(file_path)

# Example usage
file_path = 'your_existing_file.xlsx'
sheet_name = 'Sheet1'  # Change this to the desired sheet name
data_to_write = [
    {'Header 1': 'Value 1', 'Header 2': 'Value 2'},
    {'Header 1': 'Value 3', 'Header 2': 'Value 4'},
    # Add more data as needed
]

write_data_to_existing_sheet(file_path, sheet_name, data_to_write)
